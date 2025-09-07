import os
os.environ['OPENBLAS_NUM_THREADS'] = '1'   

import json
from dotenv import load_dotenv
import telebot
from datetime import datetime, timedelta
import pytz
import sqlite3
from sqlite3 import Error
import os
import pandas as pd
from openpyxl.utils import get_column_letter  
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side  
import logging
import traceback
from pathlib import Path

# Получаем текущий путь
db_path = os.path.join(os.getenv('XLSX_PATH', '/app/xlsx_reports'), 'school_bot.db')
DB_PATH = os.path.join(db_path, 'lunch_database.db')

# Токены
load_dotenv()
TOKEN = os.getenv('TOKEN')
ADMIN_TOKEN = os.getenv('ADMIN_TOKEN')
ADMIN_CHAT_ID = os.getenv('ADMIN_CHAT_ID')
XLSX_PATH = os.getenv('XLSX_PATH')
# Инициализация ботов
bot = telebot.TeleBot(TOKEN)
admin_bot = telebot.TeleBot(ADMIN_TOKEN)

def ensure_db_directory():
    db_dir = os.path.dirname(DB_PATH)
    Path(db_dir).mkdir(parents=True, exist_ok=True)
    # Установите правильные права доступа
    os.chmod(db_dir, 0o755)


def create_connection():
    try:
        conn = sqlite3.connect(DB_PATH)
        return conn
    except Error as e:
        print(f"Ошибка при подключении к БД: {e}")
        return None

def init_db():
    
    
    if not os.path.exists(DB_PATH):
        
        open(DB_PATH, 'a').close()
        print("База данных создана")

    conn = create_connection()
    if conn:
        try:
            cursor = conn.cursor()
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY,
                    chat_id INTEGER,
                    first_name TEXT,
                    last_name TEXT,
                    grade TEXT,
                    phone TEXT,
                    status TEXT,
                    user_name TEXT,
                    privil BOOLEAN,
                    last_msg INTEGER,
                    send_teacher BOOLEAN
                    
                    
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS classes (
                    class TEXT PRIMARY KEY,
                    people INTEGER
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS users_waitlist (
                    chat_id INTEGER PRIMARY KEY,
                    first_name TEXT,
                    last_name TEXT,
                    grade TEXT,
                    phone TEXT,
                    status TEXT,
                    user_name TEXT,
                    privil BOOLEAN,
                    send_req BOOLEAN
                )
            ''')
            
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS blocked_users (
                    chat_id INTEGER PRIMARY KEY,
                    user_name TEXT,
                    phone TEXT
                )
            ''')
            conn.commit()
            print("База данных и таблицы инициализированы")
        except Error as e:
            print(f"Ошибка при инициализации базы данных: {e}")
        finally:
            conn.close()

def db_check_username(username):
    conn = create_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT user_name FROM users WHERE user_name = ?", (username,))
            
            result = cursor.fetchone()
            return result is not None
            
        except Error as e:
            print(f"Ошибка при проверке ID: {e}")
            return False
        finally:
            conn.close()
    return False

def db_check_id(chat_id):
    conn = create_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT chat_id FROM users WHERE chat_id = ?", (chat_id,))
            
            result = cursor.fetchone()
            return result is not None
            
        except Error as e:
            print(f"Ошибка при проверке ID: {e}")
            return False
        finally:
            conn.close()
    return False

def db_check_status_pupil(chat_id):
    conn = create_connection()
    if conn is not None:        
        cursor = conn.cursor()
        cursor.execute("SELECT status FROM users WHERE chat_id = ?", (chat_id,))
        result = cursor.fetchone()        
        if result is not None and len(result) > 0 and result[0] == 'pupil':
            conn.close()
            return True            
        else:
            conn.close()
            
            return False

def db_check_status_teacher(chat_id):
    conn = create_connection()
    if conn is not None:        
        cursor = conn.cursor()
        cursor.execute("SELECT status FROM users WHERE chat_id = ?", (chat_id,))
        result = cursor.fetchone()
        
        if result is not None and result[0] == 'teacher':
            conn.close()
            return True            
        else:
            conn.close()
            
            return False

def db_check_status_pupil_by_username(username):
    conn = create_connection()
    if conn is not None:
        cursor = conn.cursor()
        cursor.execute("SELECT status FROM users WHERE user_name = ?", (username,))
        result = cursor.fetchone()
        
        if result[0] == 'pupil':
            conn.close()
            return True
        else:
            conn.close()
            return False
        
def db_check_status_teacher_by_username(username):
    conn = create_connection()
    if conn is not None:
        cursor = conn.cursor()
        cursor.execute("SELECT status FROM users WHERE user_name = ?", (username,))
        result = cursor.fetchone()
        
        if result[0] == 'teacher':
            conn.close()
            return True
        else:
            conn.close()
            return False
        
def db_get_lunch_info_teacher(table_name, clas):
    conn = create_connection()
    if conn is not None:
        try:
            cursor = conn.cursor()
            cursor.execute(f"SELECT name, will_eat, chat_id FROM {table_name} ORDER BY name ASC ")
            results = cursor.fetchall()
            voted_people_chat_id = []
            formatted_results = []
            for name, will_eat, chat_id in results:
                voted_people_chat_id.append(chat_id)
                status = "✅" if will_eat else "❌"
                formatted_results.append(f"{name} - {status}")
            

            cursor.execute("SELECT chat_id FROM users WHERE grade = ? AND status = ? ", (clas,'pupil'))
            all_people = cursor.fetchall()  # Сохраняем результат в переменную
            people_in_class = len(all_people)  # Используем сохраненный результат
            
            unvoted_people = []
            unvoted_names = []
            for person_unvoted in all_people:  # Используем сохраненный результат
                if person_unvoted[0] not in voted_people_chat_id:
                    unvoted_people.append(person_unvoted[0])  # Берем только chat_id из кортежа

            for unvoted_person in unvoted_people: 
                cursor.execute('SELECT first_name,last_name FROM users where chat_id =?',(unvoted_person,))
                name_tuple = cursor.fetchone()
                if name_tuple:
                    unvoted_names.append((name_tuple[1], name_tuple[0]))  # Сохраняем как кортеж (фамилия, имя)
            
            # Сортируем список по фамилии
            unvoted_names.sort()  # Сортировка по первому элементу кортежа (фамилии)
            
            # Преобразуем отсортированные кортежи в строки
            unvoted_names = [f"{last_name} {first_name}" for last_name, first_name in unvoted_names]

            cursor.execute(f"SELECT users.chat_id, users.privil, {table_name}.will_eat FROM users INNER JOIN {table_name} ON users.chat_id = {table_name}.chat_id WHERE grade = ?", (clas,))
            people = cursor.fetchall() 
            voted_people = len(people) ### 2 - льготник 3 - обеды


            eat_no = 0
            eat_yes = 0
            eat_priv = 0
            
            for person in people: 
                if person[2] == 1: 
                    if person[1] == 1: 
                        eat_priv += 1
                    elif person[1] == 0: 
                        eat_yes +=1
                else: 
                    eat_no +=1 
            
            info = ""

            info = info + f"{eat_yes} обедают\n{eat_priv} обедают как льготники \n{eat_no} не обедают" if formatted_results else "Нет данных"
            
            info = info + f"\nПроголосовало {voted_people} из {people_in_class}\n"
            
            info = info + "----------------------------------------------\n"

            info = info + "\n".join(formatted_results) 

            info = info + "\n----------------------------------------------\n"

            info = info + "Не проголосовали:\n"

            info = info + "\n".join(unvoted_names) 

            return info 
            
        except Error as e:
            tz = pytz.timezone('Asia/Yekaterinburg')
            today = datetime.now(tz)
            date_today = today.strftime("%d.%m")
            tomorrow = datetime.now(tz) + timedelta(days=1)
            date = tomorrow.strftime("%d.%m")
            create_table(date_today)
            create_table(date)
            return db_get_lunch_info_teacher(table_name, clas)
        finally:
            conn.close()
    return "Ошибка подключения к базе данных"

def db_check_username_reg(username):
    conn = create_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT user_name FROM users WHERE user_name = ?', (username,))
    info = cursor.fetchall()
    
    cursor.close()
    conn.close()
    if info is None:
        return True
    else: 
        return False 
    
def db_check_phone_reg(phone):
    conn = create_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT phone FROM users WHERE phone = ?', (phone,))
    info = cursor.fetchone()
    
    if info is None:    
        return True
    else:     
        return False 

def create_table(date):
    conn = create_connection()
    if conn is not None:
        try:
            year = datetime.now().strftime("%Y")
            table_name = f"lunch_{date.replace('.', '_')}_{year}"
            cursor = conn.cursor()
            
            cursor.execute(f''' SELECT count(name) FROM sqlite_master WHERE type='table' AND name='{table_name}' ''')
            
            if cursor.fetchone()[0] == 0:
                cursor.execute(f'''
                    CREATE TABLE IF NOT EXISTS {table_name} (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        chat_id INTEGER NOT NULL,
                        name TEXT,
                        will_eat BOOLEAN NOT NULL,
                        will_attend BOOLEAN,
                        reason TEXT
                    )
                ''')
            
            conn.commit()
            print(f"Таблица {table_name} проверена и обновлена")
            return table_name
        except Error as e:
            print(f"Ошибка при работе с таблицей: {e}")
        finally:
            conn.close()
    return None

def add_lunch_record(table_name, chat_id, will_eat, date, send = 0, will_attend=None, reason=None):
    if isinstance(will_eat, str):
        will_eat = will_eat.lower() == "да"
    if isinstance(will_attend, str) and will_attend is not None:
        will_attend = will_attend.lower() == "да"
    
    if will_eat:
        will_attend = True
        reason = None
    
    conn = create_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT first_name, last_name, grade FROM users WHERE chat_id = ?", (chat_id,))
    name_record = cursor.fetchone()
    name = f"{name_record[1]} {name_record[0]}" 
    
    
    cursor.close()

    if conn is not None:
        try:
            cursor = conn.cursor()
            
            # Проверяем существование таблицы
            cursor.execute(''' SELECT count(name) FROM sqlite_master WHERE type='table' AND name=? ''', (table_name,))
            if cursor.fetchone()[0] == 0:
                # Создаем таблицу, если она не существует
                cursor.execute(f'''
                    CREATE TABLE "{table_name}" (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        chat_id INTEGER NOT NULL,
                        name TEXT,
                        will_eat BOOLEAN NOT NULL,
                        will_attend BOOLEAN,
                        reason TEXT
                    )
                ''')
            
            # Проверяем существующую запись
            cursor.execute(f"SELECT id FROM {table_name} WHERE chat_id = ?", (chat_id,))
            existing_record = cursor.fetchone()
            
            if existing_record:
                cursor.execute(f'''
                    UPDATE {table_name}
                    SET will_eat = ?, will_attend = ?, reason = ?, name = ? 
                    WHERE chat_id = ?
                ''', (will_eat, will_attend, reason, name, chat_id, ))
            else:
                cursor.execute(f'''
                    INSERT INTO {table_name} (chat_id, name, will_eat, will_attend, reason)
                    VALUES (?, ?, ?, ?, ?)
                ''', (chat_id, name, will_eat, will_attend if will_attend is not None else False, reason))
            
           
                    
                    



            conn.commit()
            notify_teacher(table_name=table_name,date=date, chat_id = chat_id, conn = conn, cursor = cursor )

        except Error as e:
            print(f"Ошибка при добавлении записи: {e}")
        finally:
            conn.close()

def notify_teacher(table_name, date, chat_id, conn, cursor):
    
    cursor.execute("SELECT first_name, last_name, grade FROM users WHERE chat_id = ?", (chat_id,))
    name_record = cursor.fetchone()
    clas = name_record[2]
    
    # Получаем список проголосовавших
    cursor.execute(f"SELECT chat_id FROM {table_name}")
    voted_records = cursor.fetchall()
    voted_people = len(voted_records)
    voted_people_chat_id = [record[0] for record in voted_records]  # Создаем список chat_id проголосовавших
    
    cursor.execute(f"SELECT chat_id FROM users WHERE grade = ? AND status = ?", (clas, 'pupil'))
    all_people = cursor.fetchall()
    people_in_class = len(all_people)
    
    unvoted_people = []
    unvoted_names = []
    for person_unvoted in all_people:
        if person_unvoted[0] not in voted_people_chat_id:  # Теперь переменная определена
            unvoted_people.append(person_unvoted[0])

    for unvoted_person in unvoted_people: 
        cursor.execute('SELECT first_name,last_name FROM users where chat_id =?',(unvoted_person,))
        name_tuple = cursor.fetchone()
        if name_tuple:
            unvoted_names.append((name_tuple[1], name_tuple[0]))
            
    # Сортируем список по фамилии
    unvoted_names.sort()
    
    # Преобразуем отсортированные кортежи в строки
    unvoted_names = [f"{last_name} {first_name}" for last_name, first_name in unvoted_names]

    cursor.execute(f"SELECT users.chat_id, users.privil, {table_name}.will_eat FROM users INNER JOIN {table_name} ON users.chat_id = {table_name}.chat_id WHERE grade = ?", (clas,))
    people = cursor.fetchall() 
    voted_people = len(people)

    eat_no = 0
    eat_yes = 0
    eat_priv = 0
    
    for person in people: 
        if person[2] == 1: 
            if person[1] == 1: 
                eat_priv += 1
            elif person[1] == 0: 
                eat_yes +=1
        else: 
            eat_no +=1 
    
    info = ""

    info = info + f"{eat_yes} обедают\n{eat_priv} обедают как льготники \n{eat_no} не обедают"
    
    info = info + f"\nПроголосовало {voted_people} из {people_in_class}\n"
    
    info = info + "----------------------------------------------\n"

    # Здесь была ошибка - formatted_results не определен
    # Получаем отформатированные результаты голосования
    cursor.execute(f"SELECT name, will_eat FROM {table_name} WHERE chat_id IN (SELECT chat_id FROM users WHERE grade = ?)", (clas,))
    vote_results = cursor.fetchall()
    formatted_results = []
    for name, will_eat in vote_results:
        status = "✅" if will_eat else "❌"
        formatted_results.append(f"{name} - {status}")

    info = info + "\n".join(formatted_results) 

    info = info + "\n----------------------------------------------\n"

    info = info + "Не проголосовали:\n"

    info = info + "\n".join(unvoted_names) 

    return info

def get_lunch_info(chat_id, today=False):
    
    tz = pytz.timezone('Asia/Yekaterinburg')
    
    if today:
        tomorrow = datetime.now(tz)
    else:
        tomorrow = datetime.now(tz) + timedelta(days=1)
        
    # Если запрашиваемый день - воскресенье, переносим на понедельник
    if tomorrow.weekday() == 6:  # 6 = воскресенье
        tomorrow = tomorrow + timedelta(days=1)  # +1 день чтобы получить понедельник
        
    date = tomorrow.strftime("%d.%m")
    year = tomorrow.strftime("%Y")
    table_name = f"lunch_{date.replace('.', '_')}_{year}"
    
    conn = create_connection()
    if conn is not None:
        try:
            record_info = "Нет данных о голосовании"
            cursor = conn.cursor()
            
            # Проверяем существование таблицы
            cursor.execute(f"""
                SELECT name FROM sqlite_master 
                WHERE type='table' AND name='{table_name}'
            """)
            
            if not cursor.fetchone():
                # Создаем таблицу, если она не существует
                cursor.execute(f'''
                    CREATE TABLE "{table_name}" (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        chat_id INTEGER NOT NULL,
                        name TEXT NOT NULL,
                        will_eat BOOLEAN NOT NULL,
                        will_attend BOOLEAN,
                        reason TEXT
                    )
                ''')
            
            # Получаем данные пользователя
            cursor.execute(f"SELECT * FROM {table_name} WHERE chat_id = ?", (chat_id,))
            record = cursor.fetchone()
            
            if record:  # Если запись найдена
                # Получаем названия колонок
                cursor.execute(f"PRAGMA table_info({table_name})")
                columns = [column[1] for column in cursor.fetchall()]
                
                record_dict = dict(zip(columns, record))
                
                obed_info = "✅" if record_dict['will_eat'] else "❌"
                appdend_info = "✅" if record_dict['will_attend'] else "❌"
                reason_info = 'Больничный' if record_dict['reason'] == 'ill' else 'Заявление' if record_dict['reason'] == 'doc' else '-'
               
                if record_dict['will_eat']:
                    record_info = f"""
{date} 
Информация о обедах: {obed_info}
        
"""
                elif record_dict['will_attend'] == False:                 
                    record_info = f"""
{date} 
Информация о обедах: {obed_info}
Будет в школе: {appdend_info}
Причина отсутствия: {reason_info}
"""
                else:
                    record_info = f"""
{date} 
Информация о обедах: {obed_info}
Будет в школе: {appdend_info}
"""
                
            
            return record_info
            
        except Error as e:
            print (e)
        finally:
            conn.close()
    return "Ошибка подключения к базе данных"

def add_user_waitlist(new_value, column_name, chat_id):
    
    conn = create_connection()
    if conn is not None:
        try:
            cursor = conn.cursor()
            
            
            cursor.execute('SELECT chat_id FROM users_waitlist WHERE chat_id = ?', (chat_id,))
            user_exists = cursor.fetchone() is not None
            
            if user_exists:
                
                cursor.execute(f'''
                    UPDATE users_waitlist 
                    SET {column_name} = ?
                    WHERE chat_id = ?
                ''', (new_value, chat_id))
            else:
                
                cursor.execute(f'''
                    INSERT INTO users_waitlist (chat_id, {column_name})
                    VALUES (?, ?)
                ''', (chat_id, new_value))
            
            conn.commit()
            return True
                
        except Error as e:
            print(f"Ошибка при работе с базой данных: {e}")
            return False
        finally:
            conn.close()
    return False

def check_blocked_user(chat_id):
    conn = create_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT chat_id FROM blocked_users WHERE chat_id = ?", (chat_id,))
    result = cursor.fetchone()
    if result is not None:
        return True
    else:
        return False

def get_waitlist_info(chat_id):
    
    conn = create_connection()
    if conn is not None:
        try:
            cursor = conn.cursor()
            
            # Получаем данные пользователя
            cursor.execute('''
                SELECT *
                FROM users_waitlist 
                WHERE chat_id = ?
            ''', (chat_id,))
            
            result = cursor.fetchone()
            
            if result:
                return list(result)
            else:
                print(f"Пользователь с chat_id {chat_id} не найден")
                return None
                
        except Error as e:
            print(f"Ошибка при получении данных: {e}")
            return None
        finally:
            conn.close()
    return None

def get_user_info(chat_id):
    conn = create_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users WHERE chat_id = ?", (chat_id,))
    user_info = cursor.fetchone()
    return user_info

def generate_attendance_report(start_date_str, end_date_str, grade, chat_id):
    
    start_date = pd.to_datetime(start_date_str, format="%d.%m.%Y")
    end_date = pd.to_datetime(end_date_str, format="%d.%m.%Y")

    
    conn = create_connection()
    cursor = conn.cursor()

    
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tables = cursor.fetchall()

    
    records = []

    cursor.execute("SELECT first_name, last_name, chat_id FROM users WHERE grade = ? AND status = ?", (grade,'pupil'))
    
    users_id = []
    users = [{'name': f"{row[1]} {row[0]}", 'chat_id': row[2]} for row in cursor.fetchall()]  
    users.sort(key=lambda x: x['name'])  
    
    for user in users: 
        users_id.append(user['chat_id'])
    
  
    dates = []    
    records = {}
    
    for table in tables:
        attendance_dict = {}
        table_name = table[0]
        
        if table_name.startswith("lunch_"):
            try:
                date_str = table_name[6:].replace('_', '.') 
                date = pd.to_datetime(date_str, format="%d.%m.%Y") 

                if start_date <= date <= end_date:
                    dates.append(date.strftime("%d.%m.%Y"))  
                    cursor.execute(f"SELECT chat_id, will_attend, reason FROM {table_name}") 
                    attendance_data = cursor.fetchall()

                    
                    for row in attendance_data: 
                        if row[0] in users_id:  
                            user_name = next(user['name'] for user in users if user['chat_id'] == row[0])  
                            attendance_dict[user_name] = (row[1], row[2]) 
                            
            except ValueError:
                continue 
            finally:
                records[date.strftime("%d.%m.%Y")] = attendance_dict
        
   
    conn.close()
    print(records) 
  
    users_names = [user['name'] for user in users]
    report_df = pd.DataFrame(index=users_names)
    
    
    for date in dates:
        report_df[date] = ''  
        for key,info in records[date].items():
            will_attend = records[date][key][0]
            reason = records[date][key][1]
            if will_attend == 1:
                        report_df.at[key, date] = '+'
            elif will_attend == 0:
                report_df.at[key, date] = '-'                    
                if reason == 'doc':
                    report_df.at[key, date] = 'З'  
                elif reason == 'ill':
                    report_df.at[key, date] = 'Б'  
    
    file_name = f'report_{chat_id}.xlsx'
    current_directory = XLSX_PATH
    output_file = os.path.join(current_directory, 'xlsx', file_name)
    report_df.to_excel(output_file, index=True, sheet_name='Attendance Report')

    
    workbook = load_workbook(output_file)
    sheet = workbook.active

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    for i, column in enumerate(sheet.columns):
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
                cell.alignment = Alignment(horizontal='center')  
                cell.border = thin_border  
            except:
                pass
        
        
        if i == 0:
            adjusted_width = (max_length + 4)  
        else:
            adjusted_width = (max_length + 1)  
        
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    workbook.save(output_file)  
    return file_name

def get_exel_users(chat_id, clas = None ): 
    conn = create_connection()
    cursor = conn.cursor()
    if clas: 
        cursor.execute("SELECT first_name, last_name, grade, phone, user_name, privil FROM users WHERE status = ?  AND grade = ? ORDER BY grade ASC, last_name ASC ", ('pupil',clas))
    elif clas == None:
        cursor.execute("SELECT first_name, last_name, grade, phone, user_name, privil FROM users WHERE status = ? ORDER BY grade ASC, last_name ASC ", ('pupil',))

    people = cursor.fetchall()

    processed_people = []
    for person in people:
        first_name, last_name, grade, phone, user_name, privil = person
        
        if privil == 0:
            privil = '-'
        elif privil == 1:
            privil = '+'

        
        processed_people.append((first_name, last_name, grade, phone, user_name, privil))
    people = processed_people
    people_df = pd.DataFrame(data = people,
                            
                            columns = ['Имя', 'Фамилия', 'Класс', 'Телефон', 'Имя пользователя телеграмм', 'Льготнк' ] 
                             )
    
    current_directory = XLSX_PATH
    file_name = f'users_{chat_id}.xlsx'
    output_file = os.path.join(current_directory, 'xlsx', file_name)
    people_df.to_excel(output_file, index= False, sheet_name='people')

    cursor.close()
    conn.close()
    
    #Формат
    workbook = load_workbook(output_file)
    sheet = workbook.active

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))  

    for i, column in enumerate(sheet.columns):
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
                cell.alignment = Alignment(horizontal='center')  
                cell.border = thin_border  
            except:
                pass
        
        
        
        adjusted_width = (max_length + 4)  
        
            
        
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    workbook.save(output_file)
    return output_file

def get_excel_users_admin(chat_id):
    conn = create_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT first_name, last_name, grade, phone, user_name, privil, status FROM users ORDER BY status DESC, grade ASC, last_name ASC")
    people = cursor.fetchall()

    processed_people = []
    for person in people:
        first_name, last_name, grade, phone, user_name, privil, status = person
        # Заменяем privil
        if privil == 0:
            privil = '-'
        elif privil == 1:
            privil = '+'
        if status == 'pupil':
            status = 'Ученик'
        elif status == 'teacher':
            status = 'Учитель'
        
        processed_people.append((first_name, last_name, grade, phone, user_name, privil, status))
    
    people_df = pd.DataFrame(data=processed_people,
                            columns=['Имя', 'Фамилия', 'Класс', 'Телефон', 'Имя пользователя телеграмм', 'Льготнк', 'Роль'])
    
    # Создаем директорию xlsx, если она не существует
    xlsx_dir = os.path.join(XLSX_PATH, 'xlsx')
    if not os.path.exists(xlsx_dir):
        os.makedirs(xlsx_dir)

    file_name = f'users_{chat_id}.xlsx'
    output_file = os.path.join(xlsx_dir, file_name)
    people_df.to_excel(output_file, index=False, sheet_name='people')

    cursor.close()
    conn.close()
    
    # Форматирование
    workbook = load_workbook(output_file)
    sheet = workbook.active

    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    for i, column in enumerate(sheet.columns):
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            except:
                pass
        
        adjusted_width = (max_length + 4)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    workbook.save(output_file)
    return output_file

def check_req_send(chat_id):
    conn = create_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT send_req FROM users_waitlist WHERE chat_id = ?", (chat_id,))
    result = cursor.fetchone()
    if result is not None: 
        result = result[0]
    else:
        result = False
    return result 
        
def clean_req_send(chat_id):
    conn = create_connection()
    cursor = conn.cursor()
    
    
    cursor.execute("UPDATE users_waitlist SET send_req =? WHERE chat_id = ?", (False, chat_id))
    conn.commit()
    conn.close()

def generate_daily_report(grade, chat_id, today=0):
    tz = pytz.timezone('Asia/Yekaterinburg')
    current_time = datetime.now(tz)
    
    if today:  # Если нужен отчет на сегодня
        target_date = current_time
    else:  # Если нужен отчет на завтра
        target_date = current_time + timedelta(days=1)
        
    # Если целевой день - воскресенье, переносим на понедельник
    if target_date.weekday() == 6:  # 6 = воскресенье
        target_date = target_date + timedelta(days=1)
        
    date = target_date.strftime("%d_%m_%Y")
    table = f"lunch_{date}"
    conn = create_connection()
    cursor = conn.cursor()
   
    
    
    # Получаем общее количество учеников в классе
    total_people_info = cursor.execute("""
        SELECT COUNT(*) 
        FROM users 
        WHERE grade = ? AND status = 'pupil'
    """, (grade,)).fetchone()
    total_people = total_people_info[0] if total_people_info else 0

    valid_reason_people_info = cursor.execute(f"""
        SELECT t.name 
        FROM users u
        INNER JOIN {table} t ON u.chat_id = t.chat_id
        WHERE u.grade = ? AND t.reason = 'doc'
    """, (grade,)).fetchall()
    valid_reason_people = len(valid_reason_people_info)
    valid_reason_names = [name[0] for name in valid_reason_people_info]

    ill_people_info = cursor.execute(f"""
        SELECT t.name 
        FROM users u
        INNER JOIN {table} t ON u.chat_id = t.chat_id
        WHERE u.grade = ? AND t.reason = 'ill'
    """, (grade,)).fetchall()
    ill_people = len(ill_people_info)
    ill_people_names = [name[0] for name in ill_people_info]
    info = ["Класс Количество учащихся (всего)",
        "Количество отсутсвующих по болезни", 
        "ФИО обучающихся, отсутствующих по болезни",
        "Количество отсутствующих по УП",
        "ФИО обучающихся, отсутствующих по УП"]
    
    data = [total_people, 
            ill_people, 
            "\n".join([f" {name}" for name in ill_people_names]), # Используем двойной перенос строки
            valid_reason_people,
            "\n".join([f" {name}" for name in valid_reason_names])]
    df = pd.DataFrame([data],  columns=info)
    


    file_name = f'daily_report_{chat_id}.xlsx'
    current_directory = XLSX_PATH
    output_file = os.path.join(current_directory, 'xlsx', file_name)
    df.to_excel(output_file, sheet_name='Attendance Report')

    workbook = load_workbook(output_file)
    sheet = workbook.active

    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    for i, column in enumerate(sheet.columns):
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
                cell.alignment = Alignment(horizontal='center')  
                cell.border = thin_border  
            except:
                pass
        
        
        if i == 0:
            adjusted_width = (max_length + 4)  
        else:
            adjusted_width = (max_length + 1)  
        
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    workbook.save(output_file)

    return file_name

def delete_last_msg(chat_id, msg):
    conn = create_connection()
    cursor = conn.cursor()
    
    # Получаем последнее сообщение напрямую из базы
    cursor.execute("SELECT last_msg FROM users WHERE chat_id = ?", (chat_id,))
    last_msg = cursor.fetchone()
    
    if last_msg and last_msg[0]:
        try:
            bot.delete_message(chat_id=chat_id, message_id=last_msg[0])
        except:
            pass
            
    # Обновляем id последнего сообщения
    cursor.execute("UPDATE users SET last_msg = ? WHERE chat_id = ?", (msg.message_id, chat_id))
    conn.commit()
    cursor.close()
    conn.close()



def create_keyboard1():
    keyboard = telebot.types.InlineKeyboardMarkup(row_width=2)
    keyboard.add(
        telebot.types.InlineKeyboardButton("Да", callback_data="yes"),
        telebot.types.InlineKeyboardButton("Нет", callback_data="no"),
        telebot.types.InlineKeyboardButton("В главное меню", callback_data="back")
    )
    return keyboard

def create_keyboard2():
    keyboard = telebot.types.InlineKeyboardMarkup()
    keyboard.add(
        telebot.types.InlineKeyboardButton("Да", callback_data="yes_school"),
        telebot.types.InlineKeyboardButton("Нет", callback_data="no_school")
    )
    return keyboard

def create_keyboard3():
    keyboard = telebot.types.InlineKeyboardMarkup()
    keyboard.add(
        telebot.types.InlineKeyboardButton("Больничный", callback_data="ill"),
        telebot.types.InlineKeyboardButton("Заявление", callback_data="doc")
    )
    return keyboard

def create_keyboard_back():
    keyboard = telebot.types.InlineKeyboardMarkup()
    keyboard.add(telebot.types.InlineKeyboardButton("В главное меню", callback_data="back"))
    return keyboard

def create_keyboard_main():
    keyboard = telebot.types.InlineKeyboardMarkup()
    keyboard.add(telebot.types.InlineKeyboardButton("Проголосовать на завтра", callback_data="vote"))
    keyboard.add(telebot.types.InlineKeyboardButton("Информация о обедах завтра", callback_data="vote_info"))
    keyboard.add(telebot.types.InlineKeyboardButton("Информация о обедах на сегодня", callback_data="vote_info_today"))
    keyboard.add(telebot.types.InlineKeyboardButton("Мой профиль", callback_data="profile_user"))
    return keyboard

def create_keyboard_profile():
    keyboard = telebot.types.InlineKeyboardMarkup()
    keyboard.add(telebot.types.InlineKeyboardButton("Изменить данные", callback_data="edit"))
    keyboard.add(telebot.types.InlineKeyboardButton("В главное меню", callback_data="back"))
    return keyboard

def create_keyboard_phone():
    keyboard = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    button_phone = telebot.types.KeyboardButton(text="Отправить контакт", request_contact=True)
    keyboard.add(button_phone)
    return keyboard

def create_keyboard_priv():
    keyboard = telebot.types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)    
    keyboard.add(telebot.types.KeyboardButton(text="Да"))
    keyboard.add(telebot.types.KeyboardButton(text="Нет"))
    return keyboard

def create_keyboard_reg1():
    keyboard = telebot.types.InlineKeyboardMarkup()
    keyboard.add(telebot.types.InlineKeyboardButton("Зарегистрироваться", callback_data="reg"))
    return keyboard

def create_keyboard_accept_reject_block(chat_id):
    
    callback_data_accept = f'accept${chat_id}'
    callback_data_reject = f'reject${chat_id}'
    callback_data_block = f'block${chat_id}'
    callback_data_accept_teacher = f'accept_teacher${chat_id}'
    
    accept_button = telebot.types.InlineKeyboardButton(text='Принять', callback_data=callback_data_accept)
    reject_button = telebot.types.InlineKeyboardButton(text='Отклонить', callback_data=callback_data_reject)
    block_button = telebot.types.InlineKeyboardButton(text='Заблокировать', callback_data=callback_data_block)
    accept_teacher_button = telebot.types.InlineKeyboardButton(text='Принять как учителя', callback_data=callback_data_accept_teacher)

    keyboard = telebot.types.InlineKeyboardMarkup()
    keyboard.add(accept_button)
    keyboard.add(reject_button)
    keyboard.add(block_button)
    keyboard.add(accept_teacher_button)

    return keyboard

def create_keyboard_reg_check(back = False):

    keyboard = telebot.types.InlineKeyboardMarkup()
    keyboard.add(telebot.types.InlineKeyboardButton("Подтвердить", callback_data="accept"))
    keyboard.add(telebot.types.InlineKeyboardButton("Изменить", callback_data="edit"))
    if back:
        keyboard.add(telebot.types.InlineKeyboardButton("Отменить", callback_data="back"))
    return keyboard

def create_keyboard_main_teacher():
    keyboard = telebot.types.InlineKeyboardMarkup()
    keyboard.add(telebot.types.InlineKeyboardButton("Получить данные по обедам на завтра", callback_data="get_lunch_info_teacher"))
    keyboard.add(telebot.types.InlineKeyboardButton("Получить данные по обедам на сегодня", callback_data="get_lunch_info_teacher_today"))
    keyboard.add(telebot.types.InlineKeyboardButton("Получить ежедневный отчет", callback_data="get_report_daily"))
    keyboard.add(telebot.types.InlineKeyboardButton("Дополнительные функции", callback_data="main_add"))
    return keyboard

def create_keyboard_main_teacher_add():
    keyboard = telebot.types.InlineKeyboardMarkup()
    keyboard.add(telebot.types.InlineKeyboardButton("Получить данные пользователей", callback_data="get_users_exel"))    
    keyboard.add(telebot.types.InlineKeyboardButton("Получить отчет", callback_data="get_report"))
    keyboard.add(telebot.types.InlineKeyboardButton("Получить информацию о обедах других классов", callback_data="other_classes"))
    keyboard.add(telebot.types.InlineKeyboardButton("Мой профиль", callback_data="profile_user"))
    keyboard.add(telebot.types.InlineKeyboardButton("В главное меню", callback_data="back"))
    return keyboard

def create_keyboard_classes_report(): 
    keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)

    conn = create_connection()
    cursor = conn.cursor()
    cursor.execute(f"""SELECT class FROM classes""")
    classes = cursor.fetchall()
    cursor.close()
    conn.close()

    for cl in classes: 
        keyboard.add(telebot.types.InlineKeyboardButton(text=f"{cl[0]}", callback_data=f"class_report${cl[0]}"))

    keyboard.add(telebot.types.InlineKeyboardButton(text=f"Назад", callback_data=f"back"))
    return keyboard

def create_keyboard_classes_reg(): 
    keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)

    conn = create_connection()
    cursor = conn.cursor()
    cursor.execute(f"""SELECT class FROM classes""")
    classes = cursor.fetchall()
    cursor.close()
    conn.close()

    for cl in classes: 
        keyboard.add(telebot.types.InlineKeyboardButton(text=f"{cl[0]}", callback_data=f"class_reg${cl[0]}"))
    
    return keyboard

def create_keyboard_classes_edit(): 
    keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)

    conn = create_connection()
    cursor = conn.cursor()
    cursor.execute(f"""SELECT class FROM classes""")
    classes = cursor.fetchall()
    cursor.close()
    conn.close()

    for cl in classes: 
        keyboard.add(telebot.types.InlineKeyboardButton(text=f"{cl[0]}", callback_data=f"class_edit${cl[0]}"))
    
    return keyboard

def create_keyboard_classes_lunch(): 
    keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)

    conn = create_connection()
    cursor = conn.cursor()
    cursor.execute(f"""SELECT class FROM classes""")
    classes = cursor.fetchall()
    cursor.close()
    conn.close()

    for cl in classes: 
        keyboard.add(telebot.types.InlineKeyboardButton(text=f"{cl[0]}", callback_data=f"class_lunch${cl[0]}"))
    
    return keyboard

def create_keyboard_classes_daily_report(): 
    keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)

    conn = create_connection()
    cursor = conn.cursor()
    cursor.execute(f"""SELECT class FROM classes""")
    classes = cursor.fetchall()
    cursor.close()
    conn.close()
    keyboard.add(telebot.types.InlineKeyboardButton(text=f"Мой класс", callback_data=f"daily_report$myclas"))
    for cl in classes: 
        keyboard.add(telebot.types.InlineKeyboardButton(text=f"{cl[0]}", callback_data=f"daily_report${cl[0]}"))
    
    return keyboard

def create_keyboard_classes_daily_report_select(clas):
    keyboard = telebot.types.InlineKeyboardMarkup(row_width=1)
    keyboard.add(telebot.types.InlineKeyboardButton(text=f"Сегодня", callback_data=f"daily_report_select$today${clas}"))
    keyboard.add(telebot.types.InlineKeyboardButton(text=f"Завтра", callback_data=f"daily_report_select$tommorow${clas}"))

    return keyboard

def create_keyboard_classes_lunch_day(clas): 
    keyboard = telebot.types.InlineKeyboardMarkup(row_width=2)
    tz = pytz.timezone('Asia/Yekaterinburg')
    tomorrow = datetime.now(tz) + timedelta(days=1)
    today = datetime.now(tz)
    
    # Проверка на воскресенье для обоих дней
    if tomorrow.weekday() == 6:
        tomorrow = tomorrow + timedelta(days=1)
    
    date_today = today.strftime("%d_%m_%Y")
    date = tomorrow.strftime("%d_%m_%Y")    
    table_name = f"lunch_{date}"
    table_name_today = f"lunch_{date_today}"
    
    keyboard.add(telebot.types.InlineKeyboardButton(text=f"{date_today.replace('_','.')}", callback_data=f"class_day_lunch${clas}${table_name_today}$today"))
    keyboard.add(telebot.types.InlineKeyboardButton(text=f"{date.replace('_','.')}", callback_data=f"class_day_lunch${clas}${table_name}$tommorow"))
    return keyboard

@bot.message_handler(commands=['start'])
def start(message):
    print('+1')
    init_db()
    tz = pytz.timezone('Asia/Yekaterinburg')
    tomorrow = datetime.now(tz) + timedelta(days=1)
    
    # Если завтра воскресенье, переносим на понедельник
    if tomorrow.weekday() == 6:  # 6 = воскресенье
        tomorrow = tomorrow + timedelta(days=1)  # +1 день чтобы получить понедельник
    
    date = tomorrow.strftime("%d.%m")
    create_table(date)
    
    
    if db_check_id(message.chat.id):
        last_msg = get_user_info(message.chat.id)
        
        if last_msg[9]:
            try:
                bot.delete_message(chat_id=message.chat.id, message_id=last_msg[9])
            except:
                pass
        
        if db_check_status_pupil(message.chat.id):
            delete_last_msg(message.chat.id, bot.send_message(
                message.chat.id,
                f"Добрый день, {message.chat.first_name}! Выберите действие:",
                reply_markup=create_keyboard_main()
            ))
        elif db_check_status_teacher(message.chat.id):
            delete_last_msg(message.chat.id, bot.send_message(
                message.chat.id,
                f"Добрый день, {message.chat.first_name}! Выберите действие:",
                reply_markup=create_keyboard_main_teacher()
            ))
    
    elif db_check_username(f"@{message.chat.username}"):        
        chat_id = message.chat.id
        user_name = f"@{message.chat.username}"
        
        conn = create_connection()
        cursor = conn.cursor()
        cursor.execute("UPDATE users SET chat_id = ? WHERE user_name = ?", (chat_id, user_name))
        conn.commit()
        cursor.close()
        conn.close()

        last_msg = get_user_info(message.chat.id)
        
        if last_msg[9]:
            try:
                bot.delete_message(chat_id=message.chat.id, message_id=last_msg[9])
            except:
                pass
        
        if db_check_status_pupil(message.chat.id):
            delete_last_msg(message.chat.id, bot.send_message(
                message.chat.id,
                f"Добрый день, {message.chat.first_name}! Выберите действие:",
                reply_markup=create_keyboard_main()
            ))
      

    else:
        msg = bot.send_message(
            message.chat.id,
            f"Добрый день, {message.chat.first_name}! Для исползования бота, вам необходимо зарегестирвоваться",
            reply_markup=create_keyboard_reg1()
        )
    
        
    
    

# Обработчик callback-запросов
@bot.callback_query_handler(func=lambda call: True)
def callback_handler(call):
    tz = pytz.timezone('Asia/Yekaterinburg')
    tomorrow = datetime.now(tz) + timedelta(days=1)
    
    # Если завтра воскресенье, переносим на понедельник
    if tomorrow.weekday() == 6:  # 6 = воскресенье
        tomorrow = tomorrow + timedelta(days=1)  # +1 день чтобы получить понедельник
    
    today = datetime.now(tz)
    date_today = today.strftime("%d_%m_%Y")
    date = tomorrow.strftime("%d_%m_%Y")    
    table_name = f"lunch_{date}"
    table_name_today = f"lunch_{date_today}"
    
    if not db_check_id(call.message.chat.id) :
        if call.data == "reg":
            if check_blocked_user(call.message.chat.id):
                bot.edit_message_text(
                    chat_id=call.message.chat.id,
                    message_id=call.message.message_id,
                    text=f"Вы заблокированы, обратитесь к администратору")
            elif check_req_send(call.message.chat.id): 
                bot.edit_message_text(
                    chat_id=call.message.chat.id,
                    message_id=call.message.message_id,
                    text=f"Заявка уже отправленна, ожидайте подтверждения")
            elif db_check_username(f"@{call.message.chat.username}"):        
                chat_id = call.message.chat.id
                user_name = f"@{call.message.chat.username}"
                
                conn = create_connection()
                cursor = conn.cursor()
                cursor.execute("UPDATE users SET chat_id = ? WHERE user_name = ?", (chat_id, user_name))
                conn.commit()
                cursor.close()
                conn.close()
        
                
                    
                if db_check_status_pupil(call.message.chat.id):
                    msg = bot.send_message(
                        call.message.chat.id,
                        f"Добрый день, {call.message.chat.first_name}! Выберите действие:",
                        reply_markup=create_keyboard_main()
                    )
                delete_last_msg(call.message.chat.id, msg)
            elif not check_blocked_user(call.message.chat.id):
                bot.edit_message_text(
                    chat_id=call.message.chat.id,
                    message_id=call.message.message_id,
                    text=f"Введите ваше имя")               
          
                bot.register_next_step_handler(call.message, get_name)
   
        elif call.data == "edit":
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=f"Введите ваше имя",
                
            )
            bot.register_next_step_handler(call.message, get_name)
        
        elif call.data == "accept":
            winfo = get_waitlist_info(call.message.chat.id)
            tmp_msg = bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=f"Ожидайте, подтверждение. Как только администратор подтвердит ваш запрос, вы получите доступ к боту."
            )
            conn = create_connection()
            cursor = conn.cursor()
            tmp_msg = tmp_msg.message_id
            
            cursor.execute("UPDATE users SET last_msg = ?  WHERE chat_id = ?", (tmp_msg, call.message.chat.id))
            cursor.execute("UPDATE users_waitlist SET  send_req = ? WHERE chat_id = ?", (True, call.message.chat.id))
            conn.commit()
            conn.close()

            admin_bot.send_message(
                ADMIN_CHAT_ID,
                f"""Новый запрос на регистрацию
Имя: {winfo[1]}
Фамилия: {winfo[2]}
Класс: {winfo[3]}
Телефон: {winfo[4]}
Имя пользователя: {winfo[6]}
Льготник: {'Да' if winfo[7] else 'Нет' }""",
                reply_markup=create_keyboard_accept_reject_block(call.message.chat.id)
            )

        elif call.data.startswith('class_reg$'): 
            grade = call.data.split('$')[1]
            bot.edit_message_text(chat_id = call.message.chat.id,
                                  message_id = call.message.id, 
                                  text = f'Вы указали класс {grade}')    
            
            add_user_waitlist(grade, 'grade', call.message.chat.id)
            bot.send_message(
                chat_id=call.message.chat.id,
                text='Вы являетесь льготником?',
                reply_markup=create_keyboard_priv()
            )
            bot.register_next_step_handler(call.message, get_priv)
            
        else:             
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=f"Добрый день, {call.message.chat.first_name}! Для исползования бота, вам необходимо зарегестирвоваться",
                reply_markup=create_keyboard_reg1()
            )
            
    elif db_check_status_teacher(call.message.chat.id) and db_check_id(call.message.chat.id):
        
        if call.data == "get_lunch_info_teacher":
            chat_id = call.message.chat.id
            
            conn = create_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT grade FROM users WHERE chat_id = ?", (chat_id,))
            result = cursor.fetchone()  # Сохраняем результат в переменной
            clas = result[0] if result else None  # Проверяем, есть ли результат
            cursor.close()
            conn.close()

            lunch_info = db_get_lunch_info_teacher(table_name=table_name, clas=clas)
            
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=f"<b><u>Обеды {date}</u></b>\n{lunch_info}",
                reply_markup=create_keyboard_back(),
                parse_mode='HTML'
            )
        
        elif call.data == "accept":
            winfo = get_waitlist_info(call.message.chat.id)
            tmp_msg =  bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=f"Ожидайте, подтверждение. Как только администратор подтвердит ваш запрос, ваши данные изменятся.",
                reply_markup = create_keyboard_back()


            )
            admin_bot.send_message(
                ADMIN_CHAT_ID,
                f"""Новый запрос на изменение данных
Имя: {winfo[1]}
Фамилия: {winfo[2]}
Класс: {winfo[3]}
Телефон: {winfo[4]}
Имя пользователя: {winfo[6]}
Льготник: {'Да' if winfo[7] else 'Нет' }""",
                reply_markup=create_keyboard_accept_reject_block(call.message.chat.id)
            )
            conn = create_connection()
            cursor = conn.cursor()
            tmp_msg = tmp_msg.message_id
            
            cursor.execute("UPDATE users SET last_msg = ? WHERE chat_id = ?", (tmp_msg, call.message.chat.id))
            conn.commit()
            conn.close()

        elif call.data.startswith('class_reg$'): 
            grade = call.data.split('$')[1]
            bot.edit_message_text(chat_id = call.message.chat.id,
                                  message_id = call.message.id, 
                                  text = f'Вы указали класс {grade}')    
            
            add_user_waitlist(grade, 'grade', call.message.chat.id)
            bot.send_message(
                chat_id=call.message.chat.id,
                text='Вы являетесь льготником?',
                reply_markup=create_keyboard_priv()
            )
            bot.register_next_step_handler(call.message, get_priv)

        elif call.data == "edit":
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=f"Введите ваше имя",
                
            )
            bot.register_next_step_handler(call.message, get_name)

        
        
        elif call.data == "back":
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=f"Добрый день, {call.message.chat.first_name}! Выберите действие:",
                reply_markup=create_keyboard_main_teacher()
            )
        
        elif call.data == "get_lunch_info_teacher_today":
            chat_id = call.message.chat.id
            
            conn = create_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT grade FROM users WHERE chat_id = ?", (chat_id,))
            result = cursor.fetchone()  # Сохраняем результат в переменной
            clas = result[0] if result else None  # Проверяем, есть ли результат
            cursor.close()
            conn.close()

            lunch_info = db_get_lunch_info_teacher(table_name=table_name_today, clas=clas)
            
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=f"<b><u>Обеды {date_today}</u></b>\n{lunch_info}",
                reply_markup=create_keyboard_back(),
                parse_mode='HTML'
            )

        elif call.data == "get_report":
            bot.edit_message_text(chat_id = call.message.chat.id,
                                message_id = call.message.id,
                                text = "Выберите класс для отчета",
                                reply_markup = create_keyboard_classes_report())      
        
        elif call.data == "get_users_exel":  
            try:           
                 # Удаляем предыдущее сообщение
                chat_id = call.message.chat.id
                conn = create_connection()
                cursor = conn.cursor()
                cursor.execute('SELECT grade FROM users WHERE chat_id = ?', (chat_id,))
                clas = cursor.fetchone()[0]
                cursor.close()     
                conn.close()
                output_file = get_exel_users(chat_id=chat_id, clas=clas)

                with open(output_file, 'rb') as file:
                    bot.send_document(
                    chat_id=call.message.chat.id,
                    document=file
                    )


                if os.path.exists(output_file):
                    os.remove(output_file)
                
                    
                msg = bot.send_message(chat_id=call.message.chat.id,
                         text=f"Добрый день, {call.message.chat.first_name}, вы числитесь в базе данных как учитель! Выберите действие:",
                         reply_markup=create_keyboard_main_teacher())
                delete_last_msg(call.message.chat.id, msg)
                
                
            except Exception as e:
                # Запись ошибки в лог
                logging.error(f"Ошибка в блоке get_users_exel: {e}\n{traceback.format_exc()}")
          
        elif call.data.startswith('class_report$'):
            grade = call.data.split('$')[1]
            bot.edit_message_text(chat_id=call.message.chat.id,
                                  message_id = call.message.id,
                             text=f"Введите промежуток времени для отчета {grade} через тире например (09.01.2025-09.04.2025)")
            bot.register_next_step_handler(call.message, get_report, grade)
           
        elif call.data == "profile_user": 
            chat_id = call.message.chat.id
            winfo = get_user_info(chat_id)
            if winfo[6] == 'teacher':
                status = 'Учитель'
            else:
                status = 'Ученик'
            
            bot.edit_message_text (chat_id=chat_id, 
                                        message_id=call.message.message_id,
                                   text=f"""Имя: {winfo[2]}
Фамилия: {winfo[3]}
Класс: {winfo[4]}
Статус: {status}
Телефон: {winfo[5]}
Имя пользователя: {winfo[7]}
Льготник: {'Да' if winfo[8] else 'Нет' }""",
                                        reply_markup = create_keyboard_profile())

        elif call.data == "other_classes":
            bot.edit_message_text(chat_id=call.message.chat.id,
                                  message_id=call.message.id,
                                  text="Укажите класс для получения информации",
                                  reply_markup=create_keyboard_classes_lunch())

        elif call.data.startswith('class_lunch$'):
            clas = call.data.split('$')[0]
            bot.edit_message_text(chat_id=call.message.chat.id,
                                  message_id=call.message.id,
                                  text="За какой день вы ходите получить информацию?",
                                  reply_markup=create_keyboard_classes_lunch_day(clas))
            pass

        elif call.data.startswith('class_day_lunch'):
            clas = call.data.split('$')[1]
            table_name = call.data.split('$')[2]
            day = call.data.split('$')[3]
            if day == "today":
                day = date_today
            elif day == "tommorow":
                day = date
            
            lunch_info = db_get_lunch_info_teacher(table_name=table_name, clas=clas)

            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=f"<b><u>Обеды {day}</u></b>\n{lunch_info}",
                reply_markup=create_keyboard_back(),
                parse_mode='HTML'
            )
            
        elif call.data == "main_add":
            msg = bot.edit_message_text(
                chat_id = call.message.chat.id,
                message_id=call.message.id,
                text= f"Добрый день, {call.message.chat.first_name}! Выберите действие:",
                reply_markup=create_keyboard_main_teacher_add()
            )

        elif call.data == "get_report_daily":
            
            bot.edit_message_text(chat_id=call.message.chat.id,
                                message_id= call.message.id, 
                                text="Выберите класс для отчета",
                                reply_markup=create_keyboard_classes_daily_report())

        

        elif call.data.startswith("daily_report$"):
            clas = call.data.split('$')[1] 
            bot.edit_message_text(chat_id=call.message.chat.id,
                                    message_id=call.message.id,
                                    text="За какой день вы хотите получить отчет?",
                                    reply_markup=create_keyboard_classes_daily_report_select(clas))

        elif call.data.startswith("daily_report_select$"):
            info = call.data.split("$")
            clas = info[2]
            day = 1 if info[1] == "today" else 0 
            if clas == "myclas": 
                conn = create_connection()
                cursor = conn.cursor()
                cursor.execute("SELECT grade FROM users WHERE chat_id = ? ", (call.message.chat.id,))
                clas = cursor.fetchone()[0]
                cursor.close()
                conn.close()

            
            try:
                
                file_name = generate_daily_report(clas, call.message.chat.id, day )
                current_directory = XLSX_PATH
                output_file = os.path.join(current_directory, 'xlsx', file_name)

                with open(output_file, 'rb') as file:
                    bot.send_document(
                        chat_id=call.message.chat.id,
                        document=file
                    )

                if os.path.exists(output_file):
                    os.remove(output_file)

                msg = bot.send_message(chat_id=call.message.chat.id,
                                 text=f"Добрый день, {call.message.chat.first_name}, вы числитесь в базе данных как учитель! Выберите действие:",
                                 reply_markup=create_keyboard_main_teacher())
                delete_last_msg(call.message.chat.id, msg)
                
                
            except Error as e: 
                bot.send_message(chat_id=call.message.chat.id,
                                
                                 text=f"{e}",
                                 reply_markup=create_keyboard_back())
            
            


    elif db_check_id(call.message.chat.id) and db_check_status_pupil(call.message.chat.id):
        
        if call.data == "yes":
            add_lunch_record(
                table_name = table_name,
                chat_id = call.message.chat.id,
                will_eat = True,
                date = date,
                send = 1,
                will_attend=None,
                reason = None
            )
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text="Данные успешно отправлены!",
                reply_markup=create_keyboard_back()
            )

        elif call.data == "vote":
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=f"Обеды {date.replace('_','.')}",
                reply_markup=create_keyboard1()
            )

        elif call.data == "vote_info":
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=get_lunch_info(call.message.chat.id),
                reply_markup=create_keyboard_back()
            )

        elif call.data == "vote_info_today":
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=get_lunch_info(call.message.chat.id, today=True),
                reply_markup=create_keyboard_back()
            )

        elif call.data.startswith('class_reg$'): 
            grade = call.data.split('$')[1]
            bot.edit_message_text(chat_id = call.message.chat.id,
                                  message_id = call.message.id, 
                                  text = f'Вы указали класс {grade}')    
            
            add_user_waitlist(grade, 'grade', call.message.chat.id)
            bot.send_message(
                chat_id=call.message.chat.id,
                text='Вы являетесь льготником?',
                reply_markup=create_keyboard_priv()
            )
            bot.register_next_step_handler(call.message, get_priv)

        elif call.data == "edit":
          bot.edit_message_text(
              chat_id=call.message.chat.id,
              message_id=call.message.message_id,
              text=f"Введите ваше имя",
              
          )
          bot.register_next_step_handler(call.message, get_name)

        elif call.data == "accept":
            winfo = get_waitlist_info(call.message.chat.id)
            tmp_msg = bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=f"Ожидайте, подтверждение. Как только администратор подтвердит ваш запрос, ваши данные изменятся.",
                reply_markup = create_keyboard_back())
            
            
            admin_bot.send_message(
                ADMIN_CHAT_ID,
                f"""Новый запрос на изменение данных
Имя: {winfo[1]}
Фамилия: {winfo[2]}
Класс: {winfo[3]}
Телефон: {winfo[4]}
Имя пользователя: {winfo[6]}
Льготник: {'Да' if winfo[7] else 'Нет' }""",
                reply_markup=create_keyboard_accept_reject_block(call.message.chat.id)
            )
            conn = create_connection()
            cursor = conn.cursor()
            tmp_msg = tmp_msg.message_id
            
            cursor.execute("UPDATE users SET last_msg = ? WHERE chat_id = ?", (tmp_msg, call.message.chat.id))
            conn.commit()
            conn.close()

        elif call.data == "no":
            add_lunch_record(
                table_name = table_name,
                chat_id = call.message.chat.id,
                will_eat = False,
                send = 1,
                date = date,
                will_attend=None,
                reason = None
            )
          
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text="Вы будете завтра в школе?",
                reply_markup=create_keyboard2()
            )

        elif call.data == "yes_school":
            

            add_lunch_record(
                table_name = table_name,
                chat_id = call.message.chat.id,
                will_eat = False,
                date = date,
                send = 1,
                will_attend=True,
                reason = None
            )
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text="Данные успешно отправлены!",
                reply_markup=create_keyboard_back()
            )

        elif call.data == "no_school":
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text="Причина отсутствия в школе?",
                reply_markup=create_keyboard3()
            )

        elif call.data == "ill":
            

            add_lunch_record(
                table_name = table_name,
                chat_id = call.message.chat.id,
                will_eat = False,
                date = date,
                send = 1,
                will_attend = False,
                reason = "ill"
            )

            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text="Данные успешно отправлены!",
                reply_markup=create_keyboard_back()
            )

        elif call.data == "doc":
           
            add_lunch_record(
                table_name = table_name,
                chat_id = call.message.chat.id,
                will_eat = False,
                date = date,
                send = 1,
                will_attend = False,
                reason = "doc"
            )
            

            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text="Данные успешно отправлены!",
                reply_markup=create_keyboard_back()
            )

        elif call.data == "back":
            bot.edit_message_text(
                chat_id=call.message.chat.id,
                message_id=call.message.message_id,
                text=f"Добрый день, {call.message.chat.first_name}! Выберите действие:",
                reply_markup=create_keyboard_main()
            )
        
        elif call.data == "profile_user": 
            chat_id = call.message.chat.id
            winfo = get_user_info(chat_id)
            if winfo[6] == 'teacher':
                status = 'Учитель'
            else:
                status = 'Ученик'
            
            bot.edit_message_text (chat_id=chat_id, 
                                        message_id=call.message.message_id,
                                   text=f"""Имя: {winfo[2]}
Фамилия: {winfo[3]}
Класс: {winfo[4]}
Статус: {status}
Телефон: {winfo[5]}
Имя пользователя: {winfo[7]}
Льготник: {'Да' if winfo[8] else 'Нет' }""",
                                        reply_markup = create_keyboard_profile())
        

    bot.answer_callback_query(call.id)

###регистрация
def get_name(message):
    name = message.text.title()
    
    add_user_waitlist(message.chat.id, 'chat_id', message.chat.id)
    add_user_waitlist(name, 'first_name', message.chat.id)
    if message.chat.username is not None:        
        username = f"@{message.chat.username}"
        if not db_check_username_reg(username):
            add_user_waitlist(username, 'user_name', message.chat.id)
            bot.send_message(
                message.chat.id,  
                'Введите вашу фамилию', 
            )
            bot.register_next_step_handler(message, get_last_name) 
        else: 
             bot.send_message(
                message.chat.id,  
                'Пользователь с вашим username уже числится в базе данных, обратитесь к администратору', 
            )      
    else:
        bot.send_message(
            chat_id = message.chat.id,
            text = 'Для регистрации требуется username вашего телеграм аккаунта. Пожалуйста укажите его в своем профиле телеграм и попробуйте зарегестрироваться заново',
            reply_markup = create_keyboard_reg1()
        )
   
def get_last_name(message):
    last_name = message.text.title()

    add_user_waitlist(last_name, 'last_name', message.chat.id)
    bot.send_message(
        message.chat.id, 
        'Выберите ваш класс ниже', 
        reply_markup = create_keyboard_classes_reg()
    )   

def get_priv(message):
    if message.text != 'Да' and message.text != 'Нет':
        bot.send_message(
                chat_id=message.chat.id,
                text='Вы являетесь льготником? Нажмите кнопку ниже',
                reply_markup=create_keyboard_priv()
            )
        bot.register_next_step_handler(message, get_priv)
    else:        
        priv = 0
        if message.text == 'Да':
            priv = 1 
        elif message.text == 'Нет':
            priv == 0
        add_user_waitlist(priv,'privil', message.chat.id)


        remove_keyboard = telebot.types.ReplyKeyboardRemove()
        temp_msg = bot.send_message(message.chat.id, "Данные успешно сохранены!", reply_markup=remove_keyboard)
        try:
            bot.delete_message(chat_id=message.chat.id, message_id=temp_msg.message_id)
        except:
            pass
        winfo = get_waitlist_info(message.chat.id)    

        bot.send_message(
            message.chat.id, 
            'Нажмите кнопку ниже чтобы указать свой номер телефона', 
            reply_markup=create_keyboard_phone()
        )
        bot.register_next_step_handler(message, get_phone)

def get_phone(message):
    if message.contact:     
        phone = message.contact.phone_number
        chat_id = message.chat.id
        print(phone)
        print('phone', db_check_phone_reg(phone))
        db_check = db_check_id(chat_id)
        print(db_check)

        if phone.startswith('+7'):
            if len(phone) == 12:
                phone = phone
        elif  phone.startswith('7'):
            if len(phone) == 11:
                phone = f"+{phone}"
        print(phone)

        if db_check_phone_reg(phone) or db_check: 
            add_user_waitlist(phone, 'phone', message.chat.id) 

            remove_keyboard = telebot.types.ReplyKeyboardRemove()
            temp_msg = bot.send_message(message.chat.id, "Данные успешно сохранены!", reply_markup=remove_keyboard)
            try:
                bot.delete_message(chat_id=message.chat.id, message_id=temp_msg.message_id)
            except Error as e:
                print(e)
            winfo = get_waitlist_info(message.chat.id)    

            bot.send_message(
                message.chat.id,
                f"""Проверьте введенные даныные и нажмите кнопку 'Подтвердить', если все верно.
Ваши данные:
Имя: {winfo[1]}
Фамилия: {winfo[2]}
Класс: {winfo[3]}
Телефон: {winfo[4]}
Имя пользователя: {winfo[6]}
Льготник: {'Да' if winfo[7] else 'Нет' }
                """, 
                reply_markup=create_keyboard_reg_check(db_check_id(message.chat.id)))
        else: 
            remove_keyboard = telebot.types.ReplyKeyboardRemove()
            temp_msg = bot.send_message(message.chat.id, "Неверный номер", reply_markup=remove_keyboard)
            try:
                bot.delete_message(chat_id=message.chat.id, message_id=temp_msg.message_id)
            except Error as e:
                print(e)
            bot.send_message(message.chat.id, "Пользователь с данным номером уже зарегистрирован, обратитесь к администратору")
    else: 
        remove_keyboard = telebot.types.ReplyKeyboardRemove()
        temp_msg = bot.send_message(message.chat.id, "Неверно отправлен контакт! \nНажмите кнопку ниже чтобы указать свой номер телефона", reply_markup=remove_keyboard)
        try:
            bot.delete_message(chat_id=message.chat.id, message_id=temp_msg.message_id)
        except:
            pass
        bot.send_message(
        message.chat.id, 
        'Неверно отправлен контакт! \nНажмите кнопку ниже чтобы указать свой номер телефона', 
        reply_markup=create_keyboard_phone()
    )
        bot.register_next_step_handler(message, get_phone)


###создаие отчета
def get_report(message, grade): 
    try:
        dates = message.text.split('-')
        file_name = generate_attendance_report(dates[0], dates[1], grade, message.chat.id)
        current_directory = XLSX_PATH
        output_file = os.path.join(current_directory, 'xlsx', file_name)

        with open(output_file, 'rb') as file:
            bot.send_document(
                chat_id=message.chat.id,
                document=file
            )

        if os.path.exists(output_file):
            os.remove(output_file)

        msg = bot.send_message(chat_id=message.chat.id,
                         text=f"Добрый день, {message.chat.first_name}, вы числитесь в базе данных как учитель! Выберите действие:",
                         reply_markup=create_keyboard_main_teacher())
        delete_last_msg(message.chat.id, msg)
    except: 
        bot.send_message(chat_id=message.chat.id,
                         text=f"Ошибка, проверьте верность промежутка, и попробуйте заново",
                         reply_markup=create_keyboard_back())





if __name__ == '__main__':
    init_db()
    bot.polling(none_stop=True)